from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file, session
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from datetime import datetime, timedelta
import json
import os
import bcrypt
import re
import uuid
import psycopg2
from psycopg2.extras import RealDictCursor
import requests
import random
import string

app = Flask(__name__)
app.secret_key = 'supersecretkey123456'

# ===== 🌟 从请求头读取 Cookie 的钩子 =====
@app.before_request
def load_user_from_cookie_header():
    """从请求头中的 Cookie 手动加载用户，解决小程序携带 Cookie 的问题"""
    if request.endpoint in ['static', 'serve_bg']:
        return
    
    if current_user.is_authenticated:
        return
    
    cookie_header = request.headers.get('Cookie')
    if not cookie_header:
        return
    
    cookies = {}
    for item in cookie_header.split(';'):
        item = item.strip()
        if '=' in item:
            key, value = item.split('=', 1)
            cookies[key] = value
    
    if 'session' in cookies:
        try:
            from flask import session as flask_session
            if 'user_id' in flask_session:
                user_email = flask_session.get('user_id')
                if user_email:
                    user = get_user(user_email)
                    if user:
                        login_user(User(user_email))
                        print(f"✅ 通过 Cookie 自动登录: {user_email}")
        except Exception as e:
            print(f"⚠️ Cookie 加载用户失败: {e}")

# ===== 网易邮箱邮件配置 =====
NETEASE_EMAIL = "18024679346@163.com"
NETEASE_AUTH_CODE = "CNTzzMEFunsYhD8w"

def send_email(to_email, subject, body):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    try:
        smtp_server = "smtp.163.com"
        smtp_port = 465
        msg = MIMEMultipart()
        msg["From"] = f"小鱼干记账本 <{NETEASE_EMAIL}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        html_content = body.replace('\n', '<br>')
        msg.attach(MIMEText(html_content, "html", "utf-8"))
        server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        server.login(NETEASE_EMAIL, NETEASE_AUTH_CODE)
        server.sendmail(NETEASE_EMAIL, to_email, msg.as_string())
        server.quit()
        app.logger.info(f"[网易] 邮件发送成功 → {to_email}")
        return True
    except Exception as e:
        app.logger.error(f"[网易] 发送失败：{str(e)}")
        return False

# ===== 数据库连接 =====
def get_db_connection():
    database_url = os.environ.get('DATABASE_URL')
    if database_url:
        return psycopg2.connect(database_url)
    else:
        return psycopg2.connect(
            host='localhost',
            database='catbook',
            user='catbook_user',
            password=''
        )

def init_db():
    conn = get_db_connection()
    cur = conn.cursor()
    
    # 创建用户表
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            email TEXT PRIMARY KEY,
            username TEXT NOT NULL,
            password TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    ''')
    
    # 检查并添加字段
    cur.execute("SELECT column_name FROM information_schema.columns WHERE table_name='users'")
    columns = [row[0] for row in cur.fetchall()]
    
    if 'reset_token' not in columns:
        cur.execute('ALTER TABLE users ADD COLUMN reset_token TEXT')
        print("✅ 已添加 reset_token 字段")
    
    if 'reset_token_expiry' not in columns:
        cur.execute('ALTER TABLE users ADD COLUMN reset_token_expiry TEXT')
        print("✅ 已添加 reset_token_expiry 字段")
    
    if 'security_question' not in columns:
        cur.execute('ALTER TABLE users ADD COLUMN security_question TEXT')
        print("✅ 已添加 security_question 字段")
    
    if 'security_answer' not in columns:
        cur.execute('ALTER TABLE users ADD COLUMN security_answer TEXT')
        print("✅ 已添加 security_answer 字段")
    
    # ⭐ 新增：openid 字段（微信登录用）
    if 'openid' not in columns:
        cur.execute('ALTER TABLE users ADD COLUMN openid TEXT UNIQUE')
        print("✅ 已添加 openid 字段")
    
    # 创建记录表
    cur.execute('''
        CREATE TABLE IF NOT EXISTS records (
            id SERIAL PRIMARY KEY,
            date TEXT NOT NULL,
            category TEXT NOT NULL,
            amount REAL NOT NULL,
            note TEXT,
            type TEXT NOT NULL,
            created_at TEXT NOT NULL,
            user_email TEXT NOT NULL,
            FOREIGN KEY (user_email) REFERENCES users(email)
        )
    ''')
    
    conn.commit()
    cur.close()
    conn.close()
    print("✅ 数据库初始化完成！")

# ===== 用户操作 =====
def get_user(email):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM users WHERE email = %s', (email,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    return user

def get_user_by_openid(openid):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM users WHERE openid = %s', (openid,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    return user

def save_user(email, username, hashed_password, created_at, security_question=None, security_answer=None, openid=None):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        'INSERT INTO users (email, username, password, created_at, security_question, security_answer, openid) VALUES (%s, %s, %s, %s, %s, %s, %s)',
        (email, username, hashed_password, created_at, security_question, security_answer, openid)
    )
    conn.commit()
    cur.close()
    conn.close()

def update_user_token(email, token, expiry):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('UPDATE users SET reset_token = %s, reset_token_expiry = %s WHERE email = %s',
                (token, expiry, email))
    conn.commit()
    cur.close()
    conn.close()

def update_user_password(email, new_hashed_password):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('UPDATE users SET password = %s, reset_token = NULL, reset_token_expiry = NULL WHERE email = %s',
                (new_hashed_password, email))
    conn.commit()
    cur.close()
    conn.close()

def update_user_openid(email, openid):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('UPDATE users SET openid = %s WHERE email = %s',
                (openid, email))
    conn.commit()
    cur.close()
    conn.close()

# ===== 记录操作 =====
def get_records(email):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM records WHERE user_email = %s ORDER BY id DESC', (email,))
    records = cur.fetchall()
    cur.close()
    conn.close()
    return records

def add_record(email, date, category, amount, note, type_, created_at):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        'INSERT INTO records (date, category, amount, note, type, created_at, user_email) VALUES (%s, %s, %s, %s, %s, %s, %s)',
        (date, category, amount, note, type_, created_at, email)
    )
    conn.commit()
    cur.close()
    conn.close()

def delete_record(record_id, email):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM records WHERE id = %s AND user_email = %s', (record_id, email))
    conn.commit()
    cur.close()
    conn.close()

# ===== 验证 =====
def is_valid_email(email):
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

# ===== 登录管理 =====
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录哦 🐱'
login_manager.login_message_category = 'info'

class User(UserMixin):
    def __init__(self, email):
        self.id = email
        self.email = email

@login_manager.user_loader
def load_user(email):
    user = get_user(email)
    if user:
        return User(email)
    return None

# ===== 静态文件路由 =====
@app.route('/bg_pattern.png')
def serve_bg():
    return send_file('bg_pattern.png', mimetype='image/png')

# ===== 邮箱登录 =====
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        user = get_user(email)
        
        user_agent = request.headers.get('User-Agent', '')
        is_miniprogram = 'MicroMessenger' in user_agent
        
        if user:
            stored_password = user['password']
            if bcrypt.checkpw(password.encode('utf-8'), stored_password.encode('utf-8')):
                login_user(User(email))
                session['user_id'] = email
                
                if is_miniprogram:
                    resp = jsonify({'success': True, 'message': '登录成功'})
                    return resp
                else:
                    flash(f'🐱 欢迎回来，{user["username"]}！', 'success')
                    return redirect(url_for('index'))
            else:
                if is_miniprogram:
                    return jsonify({'success': False, 'message': '密码错误'}), 401
                else:
                    flash('😿 密码错误，再试一次吧', 'danger')
                    return render_template('login.html')
        else:
            if is_miniprogram:
                return jsonify({'success': False, 'message': '该邮箱未注册'}), 401
            else:
                flash('😿 该邮箱未注册，请先注册', 'warning')
                return render_template('login.html')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form.get('email')
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        security_question = request.form.get('security_question')
        security_answer = request.form.get('security_answer')
        
        if not is_valid_email(email):
            flash('😿 邮箱格式不正确', 'danger')
            return render_template('register.html')
        if not username or len(username) < 2:
            flash('😿 昵称至少2个字符', 'danger')
            return render_template('register.html')
        if not password or len(password) < 4:
            flash('😿 密码至少4个字符', 'danger')
            return render_template('register.html')
        if password != confirm_password:
            flash('😿 两次密码不一致', 'danger')
            return render_template('register.html')
        
        if get_user(email):
            flash('😿 该邮箱已被注册', 'warning')
            return render_template('register.html')
        
        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_user(email, username, hashed.decode('utf-8'), created_at, security_question, security_answer)
        
        flash('🎀 注册成功！请登录', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    flash('🐱 已退出，下次见~', 'info')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    user = get_user(current_user.email)
    username = user['username'] if user else '用户'
    return render_template('index.html', username=username, email=current_user.email)

# ===== 忘记密码 =====
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email')
        user = get_user(email)
        
        if not user:
            flash('😿 该邮箱未注册', 'danger')
            return render_template('forgot_password.html')
        
        token = str(uuid.uuid4())
        expiry = str(datetime.now().timestamp() + 900)
        update_user_token(email, token, expiry)
        
        reset_link = f"https://cat-book-62zc.onrender.com/reset_password?token={token}"
        body = f"""你好 {user['username']}，

你请求了重置密码。请点击以下链接（15分钟内有效）：

{reset_link}

如果这不是你本人的操作，请忽略此邮件。

🐱 小鱼干记账本
"""
        
        app.logger.info(f"🔍 准备发送邮件到 {email}")
        success = send_email(email, '🐱 重置你的小鱼干记账本密码', body)
        app.logger.info(f"🔍 发送结果：{success}")
        
        if success:
            flash('📧 重置邮件已发送，请查收（15分钟有效）', 'success')
        else:
            flash('😿 邮件发送失败，请稍后重试', 'danger')
        
        return redirect(url_for('login'))
    
    return render_template('forgot_password.html')

@app.route('/forgot_password_security', methods=['GET', 'POST'])
def forgot_password_security():
    email = request.args.get('email')
    if not email:
        flash('😿 请先输入邮箱', 'danger')
        return redirect(url_for('forgot_password'))
    
    user = get_user(email)
    if not user:
        flash('😿 该邮箱未注册', 'danger')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        answer = request.form.get('security_answer')
        if user['security_answer'] and user['security_answer'] == answer:
            token = str(uuid.uuid4())
            expiry = str(datetime.now().timestamp() + 900)
            update_user_token(email, token, expiry)
            flash('🎀 身份验证通过，请设置新密码', 'success')
            return redirect(url_for('reset_password', token=token))
        else:
            flash('😿 安全答案错误，请重试', 'danger')
            return render_template('forgot_password_security.html', email=email, question=user['security_question'])
    
    return render_template('forgot_password_security.html', email=email, question=user['security_question'])

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    token = request.args.get('token')
    if not token:
        flash('😿 无效的链接', 'danger')
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM users WHERE reset_token = %s', (token,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    
    if not user:
        flash('😿 无效的链接', 'danger')
        return redirect(url_for('login'))
    
    now = datetime.now().timestamp()
    if float(user['reset_token_expiry']) < now:
        flash('😿 链接已过期，请重新申请', 'danger')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        password = request.form.get('password')
        confirm = request.form.get('confirm_password')
        
        if not password or len(password) < 4:
            flash('😿 密码至少4个字符', 'danger')
            return render_template('reset_password.html', token=token)
        if password != confirm:
            flash('😿 两次密码不一致', 'danger')
            return render_template('reset_password.html', token=token)
        
        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        update_user_password(user['email'], hashed.decode('utf-8'))
        
        flash('🎀 密码重置成功！请用新密码登录', 'success')
        return redirect(url_for('login'))
    
    return render_template('reset_password.html', token=token)

# ===== 测试邮件路由 =====
@app.route('/test_email')
def test_email():
    success = send_email('fipped99@qq.com', '🐱 测试邮件', '这是一封测试邮件，如果你收到了，说明邮件功能正常！')
    if success:
        return '✅ 邮件发送成功！请检查邮箱（包括垃圾箱）'
    else:
        return '❌ 邮件发送失败，请查看 Render 日志'

# ===== ⭐ 微信登录接口 =====
# ===== ⭐ 获取用户信息接口 =====
@app.route('/api/user_info', methods=['GET'])
def api_user_info():
    """获取当前用户信息"""
    user_email = get_user_from_request()
    if not user_email:
        return jsonify({'error': '未登录'}), 401
    
    user = get_user(user_email)
    if not user:
        return jsonify({'error': '用户不存在'}), 404
    
    return jsonify({
        'email': user['email'],
        'username': user['username'],
        'has_openid': bool(user.get('openid'))
    })

# ===== ⭐ 老用户绑定微信接口 =====
@app.route('/api/bind_wechat', methods=['POST'])
def bind_wechat():
    """老用户绑定微信 openid"""
    user_email = get_user_from_request()
    if not user_email:
        return jsonify({'error': '未登录'}), 401
    
    data = request.get_json()
    openid = data.get('openid')
    
    if not openid:
        return jsonify({'success': False, 'message': '缺少openid'}), 400
    
    # 检查这个 openid 是否已被其他账号绑定
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM users WHERE openid = %s', (openid,))
    existing = cur.fetchone()
    cur.close()
    conn.close()
    
    if existing:
        return jsonify({'success': False, 'message': '该微信已绑定其他账号'}), 400
    
    # 更新当前用户的 openid
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('UPDATE users SET openid = %s WHERE email = %s', (openid, user_email))
    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({'success': True, 'message': '绑定成功'})
@app.route('/api/wx_login', methods=['POST'])
def wx_login():
    """微信登录：用 code 换 openid"""
    data = request.get_json()
    code = data.get('code')
    
    if not code:
        return jsonify({'success': False, 'message': '缺少code'}), 400
    
    # 从环境变量读取 AppSecret，更安全
    appid = 'wx815a44cbe9f4fc89'
    secret = os.environ.get('WX_SECRET', '')
    
    if not secret:
        print("❌ 错误：WX_SECRET 环境变量未设置！")
        return jsonify({'success': False, 'message': '服务器配置错误'}), 500
    
    url = f'https://api.weixin.qq.com/sns/jscode2session?appid={appid}&secret={secret}&js_code={code}&grant_type=authorization_code'
    
    try:
        resp = requests.get(url, timeout=10)
        wx_data = resp.json()
        print(f"🔍 微信返回: {wx_data}")
        
        if 'errcode' in wx_data and wx_data['errcode'] != 0:
            return jsonify({'success': False, 'message': wx_data.get('errmsg', '微信登录失败')}), 400
        
        openid = wx_data.get('openid')
        session_key = wx_data.get('session_key')
        
        if not openid:
            return jsonify({'success': False, 'message': '获取openid失败'}), 400
        
        # 查数据库：这个 openid 是否已绑定用户
        user = get_user_by_openid(openid)
        
        if user:
            # 已有用户，直接登录
            login_user(User(user['email']))
            session['user_id'] = user['email']
            return jsonify({
                'success': True,
                'message': '登录成功',
                'is_new': False,
                'user': {'email': user['email'], 'username': user['username']}
            })
        else:
            # 新用户：返回 openid，让前端引导绑定邮箱
            return jsonify({
                'success': True,
                'message': '请绑定邮箱',
                'is_new': True,
                'openid': openid
            })
            
    except requests.exceptions.Timeout:
        print("❌ 微信接口超时")
        return jsonify({'success': False, 'message': '微信接口超时，请重试'}), 500
    except Exception as e:
        print(f"❌ 微信登录异常: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ===== ⭐ 微信用户绑定邮箱接口 =====
@app.route('/api/bind_email', methods=['POST'])
def bind_email():
    """微信用户绑定邮箱"""
    data = request.get_json()
    openid = data.get('openid')
    email = data.get('email')
    username = data.get('username', '微信用户')
    
    if not openid or not email:
        return jsonify({'success': False, 'message': '缺少参数'}), 400
    
    if not is_valid_email(email):
        return jsonify({'success': False, 'message': '邮箱格式不正确'}), 400
    
    # 检查邮箱是否已被注册
    if get_user(email):
        return jsonify({'success': False, 'message': '该邮箱已被绑定，请直接登录'}), 400
    
    # 检查 openid 是否已被绑定（防止并发重复绑定）
    if get_user_by_openid(openid):
        return jsonify({'success': False, 'message': '该微信已绑定其他账号'}), 400
    
    # 生成随机密码（用户以后可以用邮箱+密码登录）
    random_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    hashed = bcrypt.hashpw(random_password.encode('utf-8'), bcrypt.gensalt())
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    try:
        save_user(email, username, hashed.decode('utf-8'), created_at, None, None, openid)
        
        # 自动登录
        user = get_user(email)
        login_user(User(email))
        session['user_id'] = email
        
        return jsonify({
            'success': True,
            'message': '绑定成功',
            'user': {'email': email, 'username': username}
        })
    except Exception as e:
        print(f"❌ 绑定失败: {e}")
        return jsonify({'success': False, 'message': '绑定失败，请重试'}), 500

# ===== 获取用户信息接口 =====
@app.route('/api/user_info', methods=['GET'])
def api_user_info():
    """获取当前用户信息"""
    user_email = get_user_from_request()
    if not user_email:
        return jsonify({'error': '未登录'}), 401
    
    user = get_user(user_email)
    if not user:
        return jsonify({'error': '用户不存在'}), 404
    
    return jsonify({
        'email': user['email'],
        'username': user['username'],
        'has_openid': bool(user.get('openid'))
    })

# ===== 从请求中获取用户 =====
def get_user_from_request():
    """从请求中获取当前用户，支持 Cookie 头"""
    if current_user.is_authenticated:
        return current_user.email
    
    cookie_header = request.headers.get('Cookie', '')
    cookies = {}
    for item in cookie_header.split(';'):
        item = item.strip()
        if '=' in item:
            key, value = item.split('=', 1)
            cookies[key] = value
    
    if 'user_id' in session:
        return session.get('user_id')
    
    return None

# ===== API 接口 =====
@app.route('/api/records', methods=['GET'])
def api_get_records():
    user_email = get_user_from_request()
    if not user_email:
        return jsonify({'error': '未登录，请先登录'}), 401
    
    try:
        records = get_records(user_email)
        return jsonify([dict(r) for r in records])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/add', methods=['POST'])
def api_add_record():
    user_email = get_user_from_request()
    if not user_email:
        return jsonify({'error': '未登录，请先登录'}), 401
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求体必须是JSON格式'}), 400
        
        date = data.get('date', datetime.now().strftime("%Y-%m-%d"))
        category = data.get('category', '其他')
        amount = float(data.get('amount', 0))
        note = data.get('note', '')
        type_ = data.get('type', '支出')
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        add_record(user_email, date, category, amount, note, type_, created_at)
        return jsonify({'success': True, 'message': '添加成功'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/delete/<int:record_id>', methods=['DELETE'])
def api_delete_record(record_id):
    user_email = get_user_from_request()
    if not user_email:
        return jsonify({'error': '未登录，请先登录'}), 401
    
    try:
        delete_record(record_id, user_email)
        return jsonify({'success': True, 'message': '删除成功'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/stats', methods=['GET'])
def api_get_stats():
    user_email = get_user_from_request()
    if not user_email:
        return jsonify({'error': '未登录，请先登录'}), 401
    
    try:
        records = get_records(user_email)
        now = datetime.now()
        month_str = f"{now.year}-{now.month:02d}"
        
        income = 0.0
        expense = 0.0
        categories = {}
        
        for r in records:
            if r['date'].startswith(month_str):
                if r['type'] == '收入':
                    income += r['amount']
                else:
                    expense += r['amount']
                    cat = r['category']
                    categories[cat] = categories.get(cat, 0.0) + r['amount']
        
        return jsonify({
            'income': income,
            'expense': expense,
            'balance': income - expense,
            'categories': categories
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/weekly_stats', methods=['GET'])
def api_get_weekly_stats():
    user_email = get_user_from_request()
    if not user_email:
        return jsonify({'error': '未登录，请先登录'}), 401
    
    try:
        records = get_records(user_email)
        now = datetime.now()
        start = now - timedelta(days=now.weekday())
        end = start + timedelta(days=6)
        
        week_records = []
        for r in records:
            if r.get('date'):
                record_date = datetime.strptime(r['date'], "%Y-%m-%d")
                if start <= record_date <= end:
                    week_records.append(r)
        
        total_income = sum(r['amount'] for r in week_records if r['type'] == '收入')
        total_expense = sum(r['amount'] for r in week_records if r['type'] == '支出')
        
        # 统计分类
        categories = {}
        for r in week_records:
            if r['type'] == '支出':
                cat = r['category']
                categories[cat] = categories.get(cat, 0.0) + r['amount']
        
        return jsonify({
            'start': start.strftime("%Y-%m-%d"),
            'end': end.strftime("%Y-%m-%d"),
            'income': total_income,
            'expense': total_expense,
            'balance': total_income - total_expense,
            'categories': categories,
            'records': week_records
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/monthly_stats', methods=['GET'])
def api_get_monthly_stats():
    user_email = get_user_from_request()
    if not user_email:
        return jsonify({'error': '未登录，请先登录'}), 401
    
    try:
        records = get_records(user_email)
        now = datetime.now()
        month_str = f"{now.year}-{now.month:02d}"
        
        income = 0.0
        expense = 0.0
        categories = {}
        
        for r in records:
            if r['date'].startswith(month_str):
                if r['type'] == '收入':
                    income += r['amount']
                else:
                    expense += r['amount']
                    cat = r['category']
                    categories[cat] = categories.get(cat, 0.0) + r['amount']
        
        return jsonify({
            'year': now.year,
            'month': now.month,
            'income': income,
            'expense': expense,
            'balance': income - expense,
            'categories': categories
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== 导出 Excel =====
@app.route('/api/export_excel', methods=['GET'])
@login_required
def export_excel():
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    records = get_records(current_user.email)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "记账本"
    
    headers = ['日期', '类别', '金额', '备注', '类型']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF8A80", end_color="FF8A80", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r['date'])
        ws.cell(row=row_idx, column=2, value=r['category'])
        ws.cell(row=row_idx, column=3, value=r['amount'])
        ws.cell(row=row_idx, column=4, value=r['note'] or '')
        ws.cell(row=row_idx, column=5, value=r['type'])
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 8
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'小鱼干记账_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

# ===== 启动 =====
if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
